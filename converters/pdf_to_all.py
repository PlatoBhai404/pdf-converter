import os
from pypdf import PdfReader
import pdfplumber
from pdf2image import convert_from_path
from pdf2docx import Converter

def extract_text_layout(pdf_path, output_txt_path):
    """Extracts text preserving visual layout."""
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text(layout=True) + "\n"
    with open(output_txt_path, "w", encoding="utf-8") as f:
        f.write(text)
    return True

def pdf_to_images(pdf_path, output_folder, poppler_path=None):
    """Converts every page of a PDF into PNG images."""
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)
    images = convert_from_path(pdf_path, poppler_path=poppler_path)
    for i, image in enumerate(images):
        image.save(os.path.join(output_folder, f"page_{i + 1}.png"), "PNG")
    return True

def pdf_to_docx(pdf_path, output_docx_path):
    """Converts PDF layout structurally into a Word Document (.docx)."""
    cv = Converter(pdf_path)
    cv.convert(output_docx_path, start=0, end=None)
    cv.close()
    return True

def pdf_to_html(pdf_path, output_html_path):
    """Converts PDF structural components into clean HTML web code."""
    cv = Converter(pdf_path)
    cv.convert(output_html_path, start=0, end=None)
    cv.close()
    return True

def pdf_to_excel(pdf_path, output_xlsx_path):
    """Extracts embedded grid tables from a PDF directly into an Excel workbook."""
    import openpyxl
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet
    
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            if tables:
                ws = wb.create_sheet(title=f"Page {i+1} Tables")
                for table in tables:
                    for row in table:
                        # Clean up None values to empty strings for Excel grid compatibility
                        ws.append([str(cell) if cell is not None else "" for cell in row])
                    ws.append([]) # Blank spacer row between distinct tables
    if len(wb.sheetnames) == 0:
        return False # No tables found to write
    wb.save(output_xlsx_path)
    return True